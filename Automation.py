# Add the User ID to the userId Excel file.
# Update the table row and table data as per the calendar date required data.

import time
from time import sleep
import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
import re
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

global email, password, otp
email = Private data
password = "Private data"
otp = Private data

# Setup Selenium WebDriver
driver = webdriver.Chrome()
driver.maximize_window()
driver.get("Private data")
driver.implicitly_wait(100)

try:
    def start_login_and_submit_otp(driver, email, password, otp):
        driver.find_element(By.XPATH, "//input[@name='username']").send_keys(email)
        driver.find_element(By.XPATH, "//input[@name='password']").send_keys(password)
        driver.find_element(By.XPATH, "//input[@value='Sign In']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//input[@id='inp']").send_keys(otp)
        driver.find_element(By.XPATH, "//input[@value='Submit']").click()
        time.sleep(5)

    def logout_login(driver, email, password, otp):
        try:
            driver.find_element(By.XPATH, "//img[@class='nameImage']").click()
            driver.find_element(By.XPATH, "//a[normalize-space()='Logout']").click()
            driver.find_element(By.XPATH, "//input[@name='username']").send_keys(email)
            driver.find_element(By.XPATH, "//input[@name='password']").send_keys(password)
            driver.find_element(By.XPATH, "//input[@value='Sign In']").click()
            time.sleep(3)
            driver.find_element(By.XPATH, "//input[@id='inp']").send_keys(otp)
            driver.find_element(By.XPATH, "//input[@value='Submit']").click()
            time.sleep(5)
        except NoSuchElementException:
            login_and_submit_otp(driver, email, password, otp)


    def login_and_submit_otp(driver, email, password, otp):
        try:
            driver.find_element(By.XPATH, "//input[@name='username']").send_keys(email)
            driver.find_element(By.XPATH, "//input[@name='password']").send_keys(password)
            driver.find_element(By.XPATH, "//input[@value='Sign In']").click()
            time.sleep(3)
            driver.find_element(By.XPATH, "//input[@id='inp']").send_keys(otp)
            driver.find_element(By.XPATH, "//input[@value='Submit']").click()
            time.sleep(5)
        except NoSuchElementException:
            driver.quit()
            driver = webdriver.Chrome()
            driver.maximize_window()
            driver.get("https://gethealthier.connectedlife.io/")
            driver.implicitly_wait(100)
            login_and_submit_otp(driver, email, password, otp)

    def Search_User(driver, username):
        try:
            time.sleep(3)
            try:
                driver.find_element(By.XPATH, "//input[@placeholder='Search']").send_keys(username)
                time.sleep(2)
                driver.find_element(By.XPATH, "//div[@class='blurry-text ng-star-inserted']").click()
            except NoSuchElementException:
                driver.refresh()
                driver.find_element(By.XPATH, "//input[@placeholder='Search']").clear()
                driver.find_element(By.XPATH, "//input[@placeholder='Search']").send_keys(username)
                time.sleep(2)
                driver.find_element(By.XPATH, "//div[@class='blurry-text ng-star-inserted']").click()
            time.sleep(2)
        except NoSuchElementException:
            logout_login(driver, email, password, otp)
            Search_User(driver, username)

    def select_date_overview(driver):
        try:
            driver.find_element(By.XPATH, "//input[@id='overview-datepicker']").click()
            # When need seleted previous month (1st line only)
            driver.find_element(By.XPATH, "//th[@class='prev available']//span").click()
            # driver.find_element(By.XPATH, "//th[@class='prev available']//span").click()
            # driver.find_element(By.XPATH, "//th[@class='prev available']//span").click()
            driver.find_element(By.XPATH, "//div[@class='drp-calendar left single']//div[@class='calendar-table']//tr[4]/td[7]").click()
            time.sleep(2)
        except NoSuchElementException:
            logout_login(driver, email, password, otp)
            Search_User(driver, username)
            select_date_overview(driver)

    def get_collect_data(driver, sheet, row):
        try:
            wait = WebDriverWait(driver, 10)
            try:
                time.sleep(3)
                hr = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='overview-card0']/div/div/div[2]/div[2]/span[1]"))).text
            except NoSuchElementException:
                print("Element not found, retrying...")
                driver.find_element(By.XPATH, "//span[normalize-space()='My Users']").click()
                time.sleep(10)
                Search_User(driver, username)
                select_date_overview(driver)
                get_collect_data(driver, sheet, row)

            if hr == "--" or hr == "__":
                hr = "Null"
            sheet.cell(row=row, column=2).value = hr

            # Sleep
            sleep_hours = driver.find_element(By.XPATH, "//*[@id='overview-card1']/div/div[1]/div[2]").text
            sleep_min = driver.find_element(By.XPATH, "//*[@id='overview-card1']/div/div[1]/div[3]").text
            if sleep_hours == "-- hr" and sleep_min == "-- min":
                total = "Null"
            else:
                total = f"{sleep_hours} {sleep_min}"
            sheet.cell(row=row, column=3).value = total

            # Step
            step = driver.find_element(By.XPATH, "//*[@id='overview-card2']/div/div[1]/div[2]/div[1]").text
            if step == "--":
                step = "Null"
            sheet.cell(row=row, column=4).value = step

            # Time Wearing Device %
            try:
                wear_time = driver.find_element(By.XPATH,"/html/body/app-root/div[3]/app-profile/div[1]/div[2]/div/app-overview/div[1]/app-battery/div/div/span[3]").text
                if wear_time == "..." or wear_time == "NA":
                    wear_time = "Null"
                wear_time = wear_time.replace("%", "").strip()
                sheet.cell(row=row, column=6).value = wear_time
            except:
                driver.find_element(By.XPATH, "//div[normalize-space()='Heart Rate']")
                sheet.cell(row=row, column=6).value = "Null"

            # Wellness Score
            wellness_score = driver.find_element(By.CSS_SELECTOR,"body > app-root:nth-child(1) > div:nth-child(3) > app-profile:nth-child(2) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > app-overview:nth-child(2) > div:nth-child(3) > div:nth-child(2) > div:nth-child(3) > app-overview-wellness-score:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > svg:nth-child(1) > text:nth-child(8) > tspan:nth-child(1)").text
            sheet.cell(row=row, column=7).value = wellness_score

            # Oxygen Saturation (%)
            oxygen_saturation = driver.find_element(By.XPATH, "//div[@class='bottom-item']//div[2]//div[2]").text
            oxygen_saturation = oxygen_saturation.replace(" %", "")
            if oxygen_saturation == "--":
                oxygen_saturation = "Null"
            sheet.cell(row=row, column=8).value = oxygen_saturation

            # Blood Pressure
            systolic = driver.find_element(By.XPATH, "//*[@id='blood-pressure-overview-dashboard']/div[3]/span[1]").text
            diastolic = driver.find_element(By.XPATH,
                                            "//*[@id='blood-pressure-overview-dashboard']/div[5]/span[1]").text
            if systolic == "--" and diastolic == "--":
                systolic = "Null"
                diastolic = "Null"
            sheet.cell(row=row, column=11).value = systolic
            sheet.cell(row=row, column=12).value = diastolic

            # Blood Glucose (mgdl/L)
            glucose = driver.find_element(By.XPATH, "//blood-glucose//div[@class='center-wrap']//div[2]/span[1]").text
            if glucose == "--" or glucose == 0:
                glucose = "Null"
            sheet.cell(row=row, column=13).value = glucose

            # HbA1c (%)
            hbA1c = driver.find_element(By.XPATH, "//*[@id='blood-glucose-dashboard']/div/div[8]/div[1]/div[2]").text
            hbA1c = hbA1c.replace(" %", "")
            if hbA1c == "--" or hbA1c == "No data":
                hbA1c == "Null"
            sheet.cell(row=row, column=14).value = hbA1c

            # Height (cms)
            height = driver.find_element(By.XPATH, "//p[@id='user-height']").text
            height = height.replace(" cm", "")
            if height == "--":
                height = "Null"
            sheet.cell(row=row, column=16).value = height

            # Cholesterol (Total)
            cholesterol = driver.find_element(By.XPATH,"//*[@id='overview-dashboards']/div[2]/div[7]/app-lipids-widget/div[1]/div/div[1]/div[1]/div[2]").text
            cholesterol = cholesterol.replace(" mmol/L", "").replace("mg/dL", "")
            if cholesterol == "--":
                cholesterol = "Null"
            sheet.cell(row=row, column=19).value = cholesterol

            #    Weight (KG) + Waist_Circumference (cms)
            try:
                driver.find_element(By.XPATH,"/html/body/app-root/div[3]/app-profile/div[1]/div[2]/div/app-overview/div[3]/div[2]/div[4]/body-shape-overview/div[1]/div/div").text
                sheet.cell(row=row, column=15).value = "Null"
                sheet.cell(row=row, column=17).value = "Null"
                sheet.cell(row=row, column=18).value = "Null"
            except:
                data1 = driver.find_element(By.XPATH, "//input[@id='overview-datepicker']").get_attribute("value")
                data2 = driver.find_element(By.XPATH, "//div[@class='seprate-name wrap-100']/div[2]").text
                days_of_week = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
                cleaned_data1 = re.sub(r"^(" + "|".join(days_of_week) + r"),\s*", "", data1)
                cleaned_data2 = data2.replace("Latest updated : ", "")
                if cleaned_data1 == cleaned_data2:
                    # Weight (KG)
                    try:
                        weight = driver.find_element(By.XPATH,"//div[@class='content-body flex-wrap ng-star-inserted']//div[@class='left-wrap']//div[@class='value']").text
                        weight = weight.replace(" kg", "")
                        if weight == "--":
                            weight == "Null"
                    except:
                        weight = "Null"
                    sheet.cell(row=row, column=15).value = weight

                    # Waist_Circumference (cms)
                    try:
                        waist_circumference = driver.find_element(By.XPATH,"//div[@class='middle-wrap']//div[@class='value']").text
                        waist_circumference = waist_circumference.replace(" cm", "")
                        if waist_circumference == "--":
                            waist_circumference == "Null"
                    except:
                        waist_circumference = driver.find_element(By.XPATH, "//div[@class='no-data-message']").text
                    sheet.cell(row=row, column=17).value = waist_circumference
                    # BMI
                    try:
                        bmi = driver.find_element(By.XPATH,"//*[@id='overview-dashboards']/div[2]/div[4]/body-shape-overview/div[1]/div/div[3]/div/div[2]").text
                        if bmi == "--":
                            bmi == "Null"
                    except:
                        bmi = driver.find_element(By.XPATH, "//div[@class='no-data-message']").text
                    sheet.cell(row=row, column=18).value = bmi
                else:
                    sheet.cell(row=row, column=15).value = "Null"
                    sheet.cell(row=row, column=17).value = "Null"
                    sheet.cell(row=row, column=18).value = "Null"

            # Health Report
            try:
                driver.find_element(By.XPATH, "//*[name()='text' and contains(@x,'3')]//*[name()='tspan']")
                report = "Null"
            except:
                report = "Yes"
            sheet.cell(row=row, column=20).value = report

        except NoSuchElementException:
            logout_login(driver, email, password, otp)
            Search_User(driver, username)
            select_date_overview(driver)
            get_collect_data(driver, sheet, row)

    def analysis_page(driver):
        try:
            driver.find_element(By.XPATH, "//a[@id='Analysis']").click()
            time.sleep(2)
        except NoSuchElementException:
            logout_login(driver, email, password, otp)
            Search_User(driver, username)
            select_date_overview(driver)
            get_collect_data(driver, sheet, row)
            analysis_page(driver)

    def select_date_analysis(driver):
        try:
            time.sleep(2)
            # Click on date picker image
            driver.find_element(By.XPATH,"//div[@class='date-picker-wrap ng-star-inserted']//label[@class='date-picker-label']//img").click()
            # Get the current month and year displayed in the date picker
            driver.find_element(By.XPATH, "//th[@class='prev available']").click()
            # driver.find_element(By.XPATH, "//th[@class='prev available']").click()
            # driver.find_element(By.XPATH, "//th[@class='prev available']").click()
            driver.find_element(By.XPATH,"//div[@class='drp-calendar right']//table[@class='table-condensed']/tbody/tr[4]/td[7]").click()
            driver.find_element(By.XPATH, "//div[@class='drp-calendar right']//table[@class='table-condensed']/tbody/tr[4]/td[7]").click()

            # Click the 'Update' button
            buttons = driver.find_elements(By.XPATH, "//body[1]/div[9]/div[4]/button[2]")
            for button in buttons:
                if button.text == "Update":
                    button.click()
            # time.sleep(1)
            # driver.find_element(By.XPATH,"//img[@class='right-arr']").click()
            time.sleep(10)
        except NoSuchElementException:
            login_and_submit_otp(driver, email, password, otp)
            Search_User(driver, username)
            select_date_overview(driver)
            get_collect_data(driver, sheet, row)
            analysis_page(driver)
            select_date_analysis(driver)

    def get_collect_data_2(driver, sheet, row):
        try:
            positive_contributions = driver.find_elements(By.XPATH, "//*[@id='wellness-score-dashboard']/div[2]/div/div/div//ul/li/div/span")
            datas = driver.find_elements(By.XPATH, "//*[@id='wellness-score-dashboard']/div[2]/div/div/div//ul/li/div/p")

            # Loop through and check for 'Total Sedentary Time'
            for positive_contribution, data in zip(positive_contributions, datas):
                positive_contribution = positive_contribution.text
                data = data.text
                if positive_contribution == 'Total Sedentary Time':
                    data = data.replace(" mins", "").replace("%", "")
                    sheet.cell(row=row, column=5).value = data
        except NoSuchElementException:
            login_and_submit_otp(driver, email, password, otp)
            Search_User(driver, username)
            select_date_overview(driver)
            get_collect_data(driver, sheet, row)
            analysis_page(driver)
            select_date_analysis(driver)
            get_collect_data_2(driver, sheet, row)

    def print_user(username):
        try:
            driver.find_element(By.XPATH, "//span[normalize-space()='My Users']").click()
            #Need To enable when using STM
            # driver.find_element(By.XPATH, "//li[@id='admin']//a[contains(text(),'Users')]").click()
            print(username)
        except NoSuchElementException:
            logout_login(driver, email, password, otp)
            Search_User(driver, username)
            select_date_overview(driver)
            get_collect_data(driver, sheet, row)
            analysis_page(driver)
            select_date_analysis(driver)
            get_collect_data_2(driver, sheet, row)
            print_user(username)

    file_path = 'user_id.xlsx'
    df = pd.read_excel(file_path)
    usernames = df['User Name'].unique()
    usernames_list = usernames.tolist()

    # Load existing workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    row = 2

    start_login_and_submit_otp(driver, email, password, otp)
    for username in usernames_list:
        Search_User(driver, username)
        select_date_overview(driver)
        get_collect_data(driver, sheet, row)
        analysis_page(driver)
        select_date_analysis(driver)
        get_collect_data_2(driver, sheet, row)
        print_user(username)
        row += 1

    workbook.save(file_path)
    print(f"File saved successfully at {file_path}")
except:
    workbook.save(file_path)
    print(f" Have same error but File saved successfully at {file_path}")